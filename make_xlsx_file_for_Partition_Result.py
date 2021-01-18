import openpyxl

wb = openpyxl.Workbook()
sheet1 = wb.active
cnt = 0
Pnum = 11
for j in range(9):
    for i in range(10):
        path = 'C:/Users/Yeonkey/Desktop/DBLAB/Instance File/Tinkered_MST inputfile/RESULT/Partition_Result/EX'+str(j+1)+'/P'+str(Pnum)+'_100000_50_'+str(i)+'_result.txt'
        f = open(path, "r")
        lines = f.readlines()
        time = lines[5][17:]
        time = time.replace(" ","")
        time = time.replace("\n","")
        length = lines[6][20:]
        length = length.replace(" ","")
        length = length.replace("\n","")
        sheet1.cell(row = 1, column = cnt+1,value = Pnum)
        sheet1.cell(row = 1, column = cnt+2,value = 'time')
        sheet1.cell(row = 1, column = cnt+3,value = 'length')
        sheet1.cell(row = i+2, column = cnt+2,value = float(time))
        sheet1.cell(row = i+2, column = cnt+3,value = float(length))


    cnt = cnt+4
    Pnum = Pnum+10
##########이상 11~91#####


cnt = 0
Pnum = 110
j = i = 0
for j in range(10):
    for i in range(10):
        path = 'C:/Users/Yeonkey/Desktop/DBLAB/Instance File/Tinkered_MST inputfile/RESULT/Partition_Result/EX'+str(j+10)+'/P'+str(Pnum)+'_100000_50_'+str(i)+'_result.txt'
        f = open(path, "r")
        lines = f.readlines()
        time = lines[5][17:]
        time = time.replace(" ","")
        time = time.replace("\n","")
        length = lines[6][20:]
        length = length.replace(" ","")
        length = length.replace("\n","")
        sheet1.cell(row = 15, column = cnt+1,value = Pnum)
        #if (j == 0):
        #    sheet1.cell(row = 15, column = cnt+1,value = 10)
        sheet1.cell(row = 15, column = cnt+2,value = 'time')
        sheet1.cell(row = 15, column = cnt+3,value = 'length')
        sheet1.cell(row = i+16, column = cnt+2,value = float(time))
        sheet1.cell(row = i+16, column = cnt+3,value = float(length))
    Pnum = Pnum+100
    cnt = cnt+4
#########이상 110~1010


cnt = 0
Pnum = 2010
j = i = 0
for j in range(4):
    for i in range(10):
        path = 'C:/Users/Yeonkey/Desktop/DBLAB/Instance File/Tinkered_MST inputfile/RESULT/Partition_Result/EX'+str(j+20)+'/P'+str(Pnum)+'_100000_50_'+str(i)+'_result.txt'
        f = open(path, "r")
        lines = f.readlines()
        time = lines[5][17:]
        time = time.replace(" ","")
        time = time.replace("\n","")
        length = lines[6][20:]
        length = length.replace(" ","")
        length = length.replace("\n","")
        sheet1.cell(row = 29, column = cnt+1,value = Pnum)
        #if (j == 0):
        #    sheet1.cell(row = 15, column = cnt+1,value = 10)
        sheet1.cell(row = 29, column = cnt+2,value = 'time')
        sheet1.cell(row = 29, column = cnt+3,value = 'length')
        sheet1.cell(row = i+30, column = cnt+2,value = float(time))
        sheet1.cell(row = i+30, column = cnt+3,value = float(length))
    Pnum = Pnum+1000
    cnt = cnt+4
wb.save('C:/Users/Yeonkey/Desktop/DBLAB/Instance File/Tinkered_MST inputfile/RESULT/Partition_Result/PARTITION_RESULT1.xlsx')

wb.close()

