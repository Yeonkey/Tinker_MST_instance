import openpyxl

wb = openpyxl.Workbook()
sheet1 = wb.active
cnt = 0
for j in range(10):
    for i in range(10):
        path = 'C:/Users/Yeonkey/Desktop/DBLAB/유전알고리즘/Tinkerd_MST inputfile/Portal_experiment_result/experiment'+str(j+1)+'/P110_100000_'+str(j+1)+'_'+str(i)+'_result.txt'
        f = open(path, "r")
        lines = f.readlines()
        time = lines[5][17:]
        time = time.replace(" ","")
        time = time.replace("\n","")
        length = lines[6][20:]
        length = length.replace(" ","")
        length = length.replace("\n","")
        sheet1.cell(row = 1, column = cnt+1,value = j)
        sheet1.cell(row = 1, column = cnt+2,value = 'time')
        sheet1.cell(row = 1, column = cnt+3,value = 'length')
        sheet1.cell(row = i+2, column = cnt+2,value = float(time))
        sheet1.cell(row = i+2, column = cnt+3,value = float(length))


    cnt = cnt+4
###################이상 0~9######################
cnt = 0
j = i = 0
for j in range(10):
    for i in range(10):
        path = 'C:/Users/Yeonkey/Desktop/DBLAB/유전알고리즘/Tinkerd_MST inputfile/Portal_experiment_result/experiment'+str(j+1)+'0/P110_100000_'+str(j+1)+'0_'+str(i)+'_result.txt'
        f = open(path, "r")
        lines = f.readlines()
        time = lines[5][17:]
        time = time.replace(" ","")
        time = time.replace("\n","")
        length = lines[6][20:]
        length = length.replace(" ","")
        length = length.replace("\n","")
        sheet1.cell(row = 15, column = cnt+1,value = j*10+10)
        if (j == 0):
            sheet1.cell(row = 15, column = cnt+1,value = 10)
        sheet1.cell(row = 15, column = cnt+2,value = 'time')
        sheet1.cell(row = 15, column = cnt+3,value = 'length')
        sheet1.cell(row = i+16, column = cnt+2,value = float(time))
        sheet1.cell(row = i+16, column = cnt+3,value = float(length))


    cnt = cnt+4
###################이상 10~100########################
cnt = 0
for i in range(10):
    path = 'C:/Users/Yeonkey/Desktop/DBLAB/유전알고리즘/Tinkerd_MST inputfile/Portal_experiment_result/experiment150/P110_100000_150_'+str(i)+'_result.txt'
    f = open(path, "r")
    lines = f.readlines()
    time = lines[5][17:]
    time = time.replace(" ","")
    time = time.replace("\n","")
    length = lines[6][20:]
    length = length.replace(" ","")
    length = length.replace("\n","")
    sheet1.cell(row = 29, column = cnt+1,value = 150)
    sheet1.cell(row = 29, column = cnt+2,value = 'time')
    sheet1.cell(row = 29, column = cnt+3,value = 'length')
    sheet1.cell(row = i+30, column = cnt+2,value = float(time))
    sheet1.cell(row = i+30, column = cnt+3,value = float(length))

cnt = cnt+4
for i in range(10):
    path = 'C:/Users/Yeonkey/Desktop/DBLAB/유전알고리즘/Tinkerd_MST inputfile/Portal_experiment_result/experiment200/P110_100000_200_'+str(i)+'_result.txt'
    f = open(path, "r")
    lines = f.readlines()
    time = lines[5][17:]
    time = time.replace(" ","")
    time = time.replace("\n","")
    length = lines[6][20:]
    length = length.replace(" ","")
    length = length.replace("\n","")
    sheet1.cell(row = 29, column = cnt+1,value = 200)
    sheet1.cell(row = 29, column = cnt+2,value = 'time')
    sheet1.cell(row = 29, column = cnt+3,value = 'length')
    sheet1.cell(row = i+30, column = cnt+2,value = float(time))
    sheet1.cell(row = i+30, column = cnt+3,value = float(length))
cnt = cnt+4

for i in range(10):
    path = 'C:/Users/Yeonkey/Desktop/DBLAB/유전알고리즘/Tinkerd_MST inputfile/Portal_experiment_result/experiment250/P110_100000_250_'+str(i)+'_result.txt'
    f = open(path, "r")
    lines = f.readlines()
    time = lines[5][17:]
    time = time.replace(" ","")
    time = time.replace("\n","")
    length = lines[6][20:]
    length = length.replace(" ","")
    length = length.replace("\n","")
    sheet1.cell(row = 29, column = cnt+1,value = 250)
    sheet1.cell(row = 29, column = cnt+2,value = 'time')
    sheet1.cell(row = 29, column = cnt+3,value = 'length')
    sheet1.cell(row = i+30, column = cnt+2,value = float(time))
    sheet1.cell(row = i+30, column = cnt+3,value = float(length))

cnt = cnt+4
for i in range(10):
    path = 'C:/Users/Yeonkey/Desktop/DBLAB/유전알고리즘/Tinkerd_MST inputfile/Portal_experiment_result/experiment300/P110_100000_300_'+str(i)+'_result.txt'
    f = open(path, "r")
    lines = f.readlines()
    time = lines[5][17:]
    time = time.replace(" ","")
    time = time.replace("\n","")
    length = lines[6][20:]
    length = length.replace(" ","")
    length = length.replace("\n","")
    sheet1.cell(row = 29, column = cnt+1,value = 300)
    sheet1.cell(row = 29, column = cnt+2,value = 'time')
    sheet1.cell(row = 29, column = cnt+3,value = 'length')
    sheet1.cell(row = i+30, column = cnt+2,value = float(time))
    sheet1.cell(row = i+30, column = cnt+3,value = float(length))

cnt = cnt+4
for i in range(10):
    path = 'C:/Users/Yeonkey/Desktop/DBLAB/유전알고리즘/Tinkerd_MST inputfile/Portal_experiment_result/experiment400/P110_100000_400_'+str(i)+'_result.txt'
    f = open(path, "r")
    lines = f.readlines()
    time = lines[5][17:]
    time = time.replace(" ","")
    time = time.replace("\n","")
    length = lines[6][20:]
    length = length.replace(" ","")
    length = length.replace("\n","")
    sheet1.cell(row = 29, column = cnt+1,value = 400)
    sheet1.cell(row = 29, column = cnt+2,value = 'time')
    sheet1.cell(row = 29, column = cnt+3,value = 'length')
    sheet1.cell(row = i+30, column = cnt+2,value = float(time))
    sheet1.cell(row = i+30, column = cnt+3,value = float(length))

cnt = cnt+4
for i in range(10):
    path = 'C:/Users/Yeonkey/Desktop/DBLAB/유전알고리즘/Tinkerd_MST inputfile/Portal_experiment_result/experiment500/P110_100000_500_'+str(i)+'_result.txt'
    f = open(path, "r")
    lines = f.readlines()
    time = lines[5][17:]
    time = time.replace(" ","")
    time = time.replace("\n","")
    length = lines[6][20:]
    length = length.replace(" ","")
    length = length.replace("\n","")
    sheet1.cell(row = 29, column = cnt+1,value = 500)
    sheet1.cell(row = 29, column = cnt+2,value = 'time')
    sheet1.cell(row = 29, column = cnt+3,value = 'length')
    sheet1.cell(row = i+30, column = cnt+2,value = float(time))
    sheet1.cell(row = i+30, column = cnt+3,value = float(length))

cnt = cnt+4
for i in range(10):
    path = 'C:/Users/Yeonkey/Desktop/DBLAB/유전알고리즘/Tinkerd_MST inputfile/Portal_experiment_result/experiment1000/P110_100000_1000_'+str(i)+'_result.txt'
    f = open(path, "r")
    lines = f.readlines()
    time = lines[5][17:]
    time = time.replace(" ","")
    time = time.replace("\n","")
    length = lines[6][20:]
    length = length.replace(" ","")
    length = length.replace("\n","")
    sheet1.cell(row = 29, column = cnt+1,value = 1000)
    sheet1.cell(row = 29, column = cnt+2,value = 'time')
    sheet1.cell(row = 29, column = cnt+3,value = 'length')
    sheet1.cell(row = i+30, column = cnt+2,value = float(time))
    sheet1.cell(row = i+30, column = cnt+3,value = float(length))

wb.save('C:/Users/Yeonkey/Desktop/DBLAB/유전알고리즘/Tinkerd_MST inputfile/Portal_experiment_result/PORTAL_RESULT.xlsx')

wb.close()

