import openpyxl

wb = openpyxl.Workbook()
sheet1 = wb.active
cnt = 0
Tnum = 10000
for j in range(10):
    for i in range(10):
        path = 'C:/Users/Yeonkey/Desktop/DBLAB/유전알고리즘/Tinkerd_MST inputfile/Terminal_Result/EX'+str(j+1)+'/P110_'+str(Tnum)+'_50_'+str(i)+'_result.txt'
        f = open(path, "r")
        lines = f.readlines()
        time = lines[5][17:]
        time = time.replace(" ","")
        time = time.replace("\n","")
        length = lines[6][20:]
        length = length.replace(" ","")
        length = length.replace("\n","")
        sheet1.cell(row = 1, column = cnt+1,value = Tnum)
        sheet1.cell(row = 1, column = cnt+2,value = 'time')
        sheet1.cell(row = 1, column = cnt+3,value = 'length')
        sheet1.cell(row = i+2, column = cnt+2,value = float(time))
        sheet1.cell(row = i+2, column = cnt+3,value = float(length))


    cnt = cnt+4
    Tnum = Tnum+10000
##########이상 10000~100000#####


cnt = 0
Tnum = 200000
j = i = 0
EXnum = 20
for j in range(4):
    for i in range(10):
        path = 'C:/Users/Yeonkey/Desktop/DBLAB/유전알고리즘/Tinkerd_MST inputfile/Terminal_Result/EX'+str(EXnum)+'/P110_'+str(Tnum)+'_50_'+str(i)+'_result.txt'
        f = open(path, "r")
        lines = f.readlines()
        time = lines[5][17:]
        time = time.replace(" ","")
        time = time.replace("\n","")
        length = lines[6][20:]
        length = length.replace(" ","")
        length = length.replace("\n","")
        sheet1.cell(row = 15, column = cnt+1,value = Tnum)
        #if (j == 0):
        #    sheet1.cell(row = 15, column = cnt+1,value = 10)
        sheet1.cell(row = 15, column = cnt+2,value = 'time')
        sheet1.cell(row = 15, column = cnt+3,value = 'length')
        sheet1.cell(row = i+16, column = cnt+2,value = float(time))
        sheet1.cell(row = i+16, column = cnt+3,value = float(length))
    Tnum = Tnum+100000
    cnt = cnt+4
    EXnum = EXnum + 10
wb.save('C:/Users/Yeonkey/Desktop/DBLAB/유전알고리즘/Tinkerd_MST inputfile/Terminal_Result/TERMINAL_RESULT.xlsx')

wb.close()

